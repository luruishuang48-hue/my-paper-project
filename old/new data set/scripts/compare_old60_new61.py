#!/usr/bin/env python3
import csv
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parents[1]
PROCESSED = BASE / "processed"
REPORTS = BASE / "reports"
OLD_PANEL = BASE.parent / "data" / "panel" / "clean_event_firm_panel.csv"
NEW_MATCHES = PROCESSED / "ai_timeline_aa_name_exact_matches.csv"

NEW_61_CSV = PROCESSED / "ai_timeline_61_event_level.csv"
PAIRS_CSV = PROCESSED / "old60_vs_ai_timeline61_objective_pairs.csv"
OLD_VIEW_CSV = PROCESSED / "old60_vs_ai_timeline61_old_view.csv"
NEW_VIEW_CSV = PROCESSED / "old60_vs_ai_timeline61_new_view.csv"
OUT_XLSX = PROCESSED / "old60_vs_ai_timeline61_comparison.xlsx"
REPORT_MD = REPORTS / "old60_vs_ai_timeline61_comparison_report.md"


def read_csv(path, encoding="utf-8-sig"):
    with path.open(encoding=encoding, newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def strip_aa_date_parentheses(value):
    return re.sub(
        r"\((Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|"
        r"January|February|March|April|May|June|July|August|September|October|November|December)"
        r"\.?\s*'?\d{2,4}\)",
        " ",
        value or "",
        flags=re.I,
    )


def signature_tokens(value):
    text = unicodedata.normalize("NFKC", strip_aa_date_parentheses(value or ""))
    text = text.lower()
    text = text.replace("‑", "-").replace("–", "-").replace("—", "-").replace("−", "-")
    text = text.replace("’", "'").replace("‘", "'")
    text = re.sub(r"\b(models?|family|announcement|updated|update|new|full)\b", " ", text)
    raw = re.findall(r"[a-z]+|\d+(?:\.\d+)?", text)
    tokens = []
    for token in raw:
        if re.fullmatch(r"\d+\.0+", token):
            token = token.split(".")[0]
        tokens.append(token)
    return tuple(sorted(tokens))


def split_names(value):
    return [part.strip() for part in re.split(r"\s*;\s*", value or "") if part.strip()]


def name_signatures(names):
    out = {}
    for name in names:
        sig = signature_tokens(name)
        if sig:
            out[sig] = name
    return out


def year_month_from_date(value):
    match = re.match(r"^(\d{4})[-/](\d{1,2})", value or "")
    if not match:
        return ""
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def month_gap(left, right):
    if not left or not right:
        return ""
    left_year, left_month = map(int, left.split("-"))
    right_year, right_month = map(int, right.split("-"))
    return abs((left_year - right_year) * 12 + (left_month - right_month))


def build_new_events():
    rows = read_csv(NEW_MATCHES)
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["event_id"]].append(row)
    out = []
    for event_id, group in grouped.items():
        first = group[0]
        model_names = [row["ai_model_name"] for row in group]
        aa_names = [row["aa_name"] for row in group]
        aa_dates = [row["aa_release_date"] for row in group if row.get("aa_release_date")]
        creators = sorted({row["aa_creator"] for row in group if row.get("aa_creator")})
        out.append(
            {
                "new_event_id": event_id,
                "new_ai_month": first["ai_year_month"],
                "new_ai_year": first["ai_year"],
                "new_model_names": "; ".join(model_names),
                "new_aa_names": "; ".join(aa_names),
                "new_aa_release_dates": "; ".join(aa_dates),
                "new_aa_creators": "; ".join(creators),
                "new_model_count": len(group),
                "new_event_text": first["event_text"],
                "new_signatures": name_signatures(model_names + aa_names),
            }
        )
    return sorted(out, key=lambda row: (row["new_ai_month"], row["new_event_id"]))


def build_old_events():
    rows = read_csv(OLD_PANEL)
    out = []
    seen = set()
    for row in rows:
        event_id = row["final_event_id"]
        if event_id in seen:
            continue
        seen.add(event_id)
        names = (
            split_names(row.get("model_names"))
            + split_names(row.get("aa_model_names"))
            + [
                row.get("event_name", ""),
                row.get("representative_aa_model_name", ""),
            ]
        )
        out.append(
            {
                "old_event_id": event_id,
                "old_release_date": row.get("release_date", ""),
                "old_release_month": year_month_from_date(row.get("release_date", "")),
                "old_event_name": row.get("event_name", ""),
                "old_model_names": row.get("model_names", ""),
                "old_aa_names": row.get("aa_model_names", ""),
                "old_creator": row.get("true_model_creator", ""),
                "old_model_family": row.get("model_family", ""),
                "old_signatures": name_signatures(names),
            }
        )
    return sorted(out, key=lambda row: (row["old_release_date"], row["old_event_id"]))


def classify_gap(gap):
    if gap == "":
        return "missing_date"
    if gap == 0:
        return "same_month"
    if gap == 1:
        return "adjacent_month"
    return "time_far"


def build_pairs(new_events, old_events):
    pairs = []
    for new_row in new_events:
        for old_row in old_events:
            overlap = set(new_row["new_signatures"]) & set(old_row["old_signatures"])
            if not overlap:
                continue
            labels = []
            for sig in sorted(overlap):
                labels.append(f"{new_row['new_signatures'][sig]} = {old_row['old_signatures'][sig]}")
            gap = month_gap(new_row["new_ai_month"], old_row["old_release_month"])
            pairs.append(
                {
                    "match_timing": classify_gap(gap),
                    "month_gap": gap,
                    "matched_names": "; ".join(labels),
                    "new_event_id": new_row["new_event_id"],
                    "new_ai_month": new_row["new_ai_month"],
                    "new_model_names": new_row["new_model_names"],
                    "new_aa_names": new_row["new_aa_names"],
                    "new_event_text": new_row["new_event_text"],
                    "old_event_id": old_row["old_event_id"],
                    "old_release_date": old_row["old_release_date"],
                    "old_event_name": old_row["old_event_name"],
                    "old_model_names": old_row["old_model_names"],
                    "old_aa_names": old_row["old_aa_names"],
                    "old_creator": old_row["old_creator"],
                }
            )
    return sorted(
        pairs,
        key=lambda row: (
            row["new_ai_month"],
            row["new_event_id"],
            row["old_release_date"],
            row["old_event_id"],
        ),
    )


def join_values(rows, key):
    return "; ".join(str(row[key]) for row in rows if row.get(key) not in {"", None})


def build_views(new_events, old_events, pairs):
    pairs_by_new = defaultdict(list)
    pairs_by_old = defaultdict(list)
    for pair in pairs:
        pairs_by_new[pair["new_event_id"]].append(pair)
        pairs_by_old[pair["old_event_id"]].append(pair)

    new_view = []
    for row in new_events:
        matched = pairs_by_new.get(row["new_event_id"], [])
        new_view.append(
            {
                "new_match_status": "matched_old60" if matched else "new61_only",
                "new_event_id": row["new_event_id"],
                "new_ai_month": row["new_ai_month"],
                "new_model_names": row["new_model_names"],
                "new_aa_names": row["new_aa_names"],
                "new_aa_release_dates": row["new_aa_release_dates"],
                "matched_old_event_ids": join_values(matched, "old_event_id"),
                "matched_old_dates": join_values(matched, "old_release_date"),
                "matched_old_event_names": join_values(matched, "old_event_name"),
                "match_timing": join_values(matched, "match_timing"),
                "month_gaps": join_values(matched, "month_gap"),
                "matched_names": join_values(matched, "matched_names"),
                "new_event_text": row["new_event_text"],
            }
        )

    old_view = []
    for row in old_events:
        matched = pairs_by_old.get(row["old_event_id"], [])
        old_view.append(
            {
                "old_match_status": "matched_new61" if matched else "old60_only",
                "old_event_id": row["old_event_id"],
                "old_release_date": row["old_release_date"],
                "old_event_name": row["old_event_name"],
                "old_model_names": row["old_model_names"],
                "old_aa_names": row["old_aa_names"],
                "old_creator": row["old_creator"],
                "matched_new_event_ids": join_values(matched, "new_event_id"),
                "matched_new_months": join_values(matched, "new_ai_month"),
                "matched_new_model_names": join_values(matched, "new_model_names"),
                "match_timing": join_values(matched, "match_timing"),
                "month_gaps": join_values(matched, "month_gap"),
                "matched_names": join_values(matched, "matched_names"),
            }
        )
    return new_view, old_view


def write_workbook(summary_rows, pairs, old_view, new_view, old_events, new_events):
    wb = Workbook()
    sheets = [
        ("summary", summary_rows),
        ("objective_pairs", pairs),
        ("old60_view", old_view),
        ("new61_view", new_view),
        (
            "old60_events",
            [
                {k: v for k, v in row.items() if k != "old_signatures"}
                for row in old_events
            ],
        ),
        (
            "new61_events",
            [
                {k: v for k, v in row.items() if k != "new_signatures"}
                for row in new_events
            ],
        ),
    ]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", color="000000")
    for idx, (name, rows) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = name
        fields = list(rows[0].keys()) if rows else ["empty"]
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_idx, field in enumerate(fields, start=1):
            width = min(max(len(field) + 2, 14), 48)
            for cell in ws[get_column_letter(column_idx)][1:30]:
                width = min(max(width, min(len(str(cell.value or "")) + 2, 70)), 70)
            ws.column_dimensions[get_column_letter(column_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    wb.save(OUT_XLSX)


def main():
    new_events = build_new_events()
    old_events = build_old_events()
    pairs = build_pairs(new_events, old_events)
    new_view, old_view = build_views(new_events, old_events, pairs)

    matched_new = {row["new_event_id"] for row in pairs}
    matched_old = {row["old_event_id"] for row in pairs}
    timing_counts = Counter(row["match_timing"] for row in pairs)

    summary_rows = [
        {"item": "new_61_events", "value": len(new_events)},
        {"item": "old_60_events", "value": len(old_events)},
        {"item": "objective_name_pairs", "value": len(pairs)},
        {"item": "new_events_with_any_old_match", "value": len(matched_new)},
        {"item": "old_events_with_any_new_match", "value": len(matched_old)},
        {"item": "new61_only_events", "value": len(new_events) - len(matched_new)},
        {"item": "old60_only_events", "value": len(old_events) - len(matched_old)},
        {"item": "same_month_pairs", "value": timing_counts["same_month"]},
        {"item": "adjacent_month_pairs", "value": timing_counts["adjacent_month"]},
        {"item": "time_far_pairs", "value": timing_counts["time_far"]},
    ]

    new_fields = [
        "new_event_id",
        "new_ai_month",
        "new_ai_year",
        "new_model_names",
        "new_aa_names",
        "new_aa_release_dates",
        "new_aa_creators",
        "new_model_count",
        "new_event_text",
    ]
    write_csv(
        NEW_61_CSV,
        [{field: row.get(field, "") for field in new_fields} for row in new_events],
        new_fields,
    )
    write_csv(PAIRS_CSV, pairs, list(pairs[0].keys()) if pairs else ["empty"])
    write_csv(OLD_VIEW_CSV, old_view, list(old_view[0].keys()) if old_view else ["empty"])
    write_csv(NEW_VIEW_CSV, new_view, list(new_view[0].keys()) if new_view else ["empty"])
    write_workbook(summary_rows, pairs, old_view, new_view, old_events, new_events)

    lines = [
        "# Old 60 与 AI Timeline 61 逐条对照",
        "",
        "规则。",
        "",
        "- 新表是 76 条名字严格匹配行按 AI Timeline `event_id` 合并后的 61 条。",
        "- 旧表来自 `data/panel/clean_event_firm_panel.csv` 中 60 个唯一 `final_event_id`。",
        "- 只要旧表和新表有任一模型名或 AA 模型名的规范化签名完全一致，就记为一组客观对应。",
        "- 规范化只处理大小写、标点、连字符、月份括号和 `model/family/new/full` 等通用词。",
        "",
        "## 数量",
        "",
    ]
    for row in summary_rows:
        lines.append(f"- {row['item']} {row['value']}")
    lines.extend(
        [
            "",
            "## 结论",
            "",
            "这两张表不是逐条一一对应。它们有一个明显重叠的核心，但也各自有大量独有条目。",
            "",
            "## 输出",
            "",
            f"- `{OUT_XLSX}`",
            f"- `{PAIRS_CSV}`",
            f"- `{OLD_VIEW_CSV}`",
            f"- `{NEW_VIEW_CSV}`",
            f"- `{NEW_61_CSV}`",
        ]
    )
    REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"new_events={len(new_events)}")
    print(f"old_events={len(old_events)}")
    print(f"objective_pairs={len(pairs)}")
    print(f"new_matched={len(matched_new)}")
    print(f"old_matched={len(matched_old)}")
    print(f"new_only={len(new_events) - len(matched_new)}")
    print(f"old_only={len(old_events) - len(matched_old)}")
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
